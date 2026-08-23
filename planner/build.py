#!/usr/bin/env python3
"""Build the ICT257 study planner spreadsheet.

Run from the repository root:

    python3 planner/build.py

The script is self contained. It writes planner/ict257-study-planner.xlsx and
overwrites any existing copy. Every date in the workbook is an Excel formula
driven by one input cell, so nobody has to hand edit the binary next semester.
"""

from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------
# Palette, taken from the charts already in this repository.
# --------------------------------------------------------------------------
NAVY = "051C2C"
BLUE_1 = "034B6F"
BLUE_2 = "0679A8"
BLUE_3 = "2AA3D6"
BLUE_4 = "9CC7E0"
AMBER = "E5A000"
QUIET = "C9D2D8"
RULE = "B3B3B3"
GRID = "E4E8EB"
SUBTITLE = "5A6B75"
WHITE = "FFFFFF"

FONT_NAME = "Calibri"

# --------------------------------------------------------------------------
# Source data.
# --------------------------------------------------------------------------
CHAPTERS = {
    "RH124": {
        0: "Preface - Introduction",
        1: "Introduction to Red Hat Enterprise Linux",
        2: "Accessing the Command Line",
        3: "Getting Help from Local Documentation",
        4: "Registering Systems for Red Hat Support",
        5: "Getting AI-assisted Help with Red Hat Enterprise Linux Lightspeed",
        6: "Navigating the File-system Hierarchy",
        7: "Managing Files from the Command Line",
        8: "Editing Text Files",
        9: "Redirecting Shell Input and Output",
        10: "Managing Local Users and Groups",
        11: "Controlling Access to Files",
        12: "Installing and Updating Software with RPM",
        13: "Installing and Updating Applications by Using Flatpak",
        14: "Accessing Removable Media",
        15: "Monitoring and Managing Linux Processes",
        16: "Controlling Services and Daemons",
        17: "Introduction to Networking",
        18: "Managing Network Configuration",
        19: "Configuring and Securing SSH",
        20: "Comprehensive Review",
    },
    "RH134": {
        0: "Preface - Introduction",
        1: "Shell Scripting and the Command Line",
        2: "Using Regular Expressions for Practical Applications",
        3: "Scheduling User Tasks",
        4: "Scheduling System Tasks",
        5: "Analyzing and Storing Logs",
        6: "Managing Security with SELinux",
        7: "Archiving Files",
        8: "Transferring Files",
        9: "Tuning System Performance",
        10: "Managing Basic Storage",
        11: "Managing Storage with Logical Volume Manager",
        12: "Controlling and Troubleshooting the Boot Process",
        13: "Recovering Superuser Access",
        14: "Managing Network Security",
        15: "Accessing Network-attached Storage",
        16: "Installing Red Hat Enterprise Linux",
        17: "Managing Containers with Podman",
        18: "Working with Image-based Red Hat Enterprise Linux",
        19: "Comprehensive Review",
    },
}

# Teaching week -> (course, chapter numbers, focus). Taken from lessons.md.
WEEKS = {
    1: ("RH124", [0, 1, 2, 5], "Getting oriented, the command-line assistant and the shell"),
    2: ("RH124", [3, 4, 6, 7], "Manual pages, registering systems, the file-system hierarchy and working with files"),
    3: ("RH124", [8, 9, 10], "Editing text, redirecting output, users and groups"),
    4: ("RH124", [11, 12, 13], "File permissions, RPM packages and Flatpak applications"),
    5: ("RH124", [14, 15, 16], "File systems, locating files, processes and system services"),
    6: ("RH124", [17, 18, 19], "Network addressing and configuration, name resolution and secure remote access"),
    7: ("RH134", [1, 2, 3], "Shell scripts, regular expressions and scheduled jobs"),
    8: ("RH134", [4, 5], "Recurring jobs, logs, journals and keeping time"),
    9: ("RH134", [6, 7, 8, 9], "SELinux, archives, secure file transfer and tuning profiles"),
    10: ("RH134", [10, 11, 12], "Partitions, swap, logical volumes and booting"),
    11: ("RH134", [13, 14, 15, 16], "Boot troubleshooting, firewalls, network file systems and installation"),
    12: ("RH134", [17, 18], "Containers and image mode"),
}

# lessons.md schedules no new chapters after week 12. Week 13 is catch-up and
# week 14 is revision, and that is where the two comprehensive reviews go.
EXTRA_WEEKS = {
    13: "Catch-up week. No new chapters. Overflow from the twelve teaching weeks.",
    14: "Revision week. Objective by objective, with both comprehensive reviews open to you.",
}

# Chapters lessons.md does not place in a teaching week.
UNSCHEDULED = {
    ("RH124", 20): 13,
    ("RH134", 0): 7,
    ("RH134", 19): 14,
}

NOTES = {
    ("RH124", 0): "Sections 00.02 and 00.03 only. They cover the lab environment and the exercise types.",
    ("RH124", 1): "Context, not a skill. No objective matches it.",
    ("RH124", 5): "Brought forward into week 1 so you have the assistant from the start. You will not have it in the exam.",
    ("RH124", 20): "Five review labs. Open to you from week 6, so start earlier if you can.",
    ("RH134", 0): "Not scheduled. It orients you to a self-paced Red Hat course. Skim it and move on.",
    ("RH134", 6): "Hardest third of the module starts here. Weeks 9 to 11 deserve the most revision time.",
    ("RH134", 16): "No objective matches it. Cover the chapters the exam asks for first.",
    ("RH134", 17): "No objective matches it. Cover the chapters the exam asks for first.",
    ("RH134", 18): "No objective matches it. Cover the chapters the exam asks for first.",
    ("RH134", 19): "Labs 19.02 to 19.04 are open from week 11. Lab 19.05 rests on week 12 containers.",
}

# Optional practice challenges, from practice.md. Sorted by the week each one
# becomes possible.
PRACTICE = [
    (1, "The pool of tears, and another beside it", 3, "Doable", "RHCSA-1.2, 1.1", "nothing"),
    (2, "A small door, and one sentence from the Queen", 4, "Moderate", "RHCSA-9.4, 9.1, 9.3, 9.2", "lab start users-password"),
    (3, "Everybody has won, and all must have prizes", 4, "Doable", "RHCSA-10.2, 1.10", "lab start perms-default"),
    (4, "The bottle marked drink me, and where it came from", 4, "Moderate", "RHCSA-2.1, 2.2, 7.5", "lab start software-dnf"),
    (5, "A cake marked eat me, for one guest only", 4, "Doable", "RHCSA-2.3, 2.4", "lab start flatpak-configure"),
    (6, "Clean cup, move down", 5, "Moderate", "RHCSA-6.5, 1.8, 1.9, 1.10", "lab start perms-cli"),
    (7, "The White Rabbit must not be late", 5, "Doable", "RHCSA-7.2, 4.9, 8.3", "lab start services-identify"),
    (8, "Who are you, said the Caterpillar", 6, "Doable", "RHCSA-10.3, 1.4", "lab start ssh-keyauth"),
    (9, "Two, Five and Seven paint the roses", 7, "Moderate", "RHCSA-3.3, 3.1, 3.2, 3.4", "lab start scripts-loops"),
    (10, "One card in the whole pack", 7, "Doable", "RHCSA-1.3, 1.7, 1.11", "lab start regexes-regex"),
    (11, "A watch that tells the day of the month", 8, "Moderate", "RHCSA-7.1 (at and cron only), 1.2", "lab start scheduling-cron"),
    (12, "The jury writes it all down", 8, "Moderate", "RHCSA-4.7, 4.8, 7.4", "lab start logs-systemd"),
    (13, "A caucus race with no winner", 9, "Moderate", "RHCSA-4.4, 4.5, 4.6", "lab start tuning-nice"),
    (14, "The Cheshire Cat fades, but the grin stays", 9, "Moderate", "RHCSA-10.4, 10.5, 10.8", "lab start selinux-booleans"),
    (15, "Alice grows, and the bottle is put away", 10, "Hard", "RHCSA-6.4, 5.1, 5.2, 5.3", "lab start lvm-extend"),
    (16, "No room, said the Hare, and there was plenty", 10, "Hard", "RHCSA-5.4, 5.2, 5.3, 5.6", "lab start lvm-create"),
    (17, "The treacle well, and what was drawn from it", 10, "Hard", "RHCSA-6.1 (XFS and ext4 only), 1.6, 4.10, 5.5", "lab start archive-manage"),
    (18, "A door that opens only when asked", 11, "Doable", "RHCSA-6.2, 6.3, 4.9", "lab start nfsclient-autofs"),
    (19, "The flamingo lent for the game, and given back", 11, "Moderate", "RHCSA-10.7, 10.1", "lab start netsecurity-ports"),
    (20, "What is the use of a book without pictures", 11, "Hard", "RHCSA-5.1, 5.5, 10.6, 10.1", "lab start storage-partitions"),
    (21, "A passage of their own to the garden", 11, "Moderate", "RHCSA-8.1, 8.2, 8.4", "lab start net-edit"),
    (22, "Begin at the beginning, said the King", 11, "Hard", "RHCSA-7.6, 4.1, 4.2, 4.3, 7.3, 1.5", "lab start boot-grub"),
]

DEFAULT_START = dt.date(2026, 8, 3)  # Monday of teaching week 1. Change in the sheet, not here.

STATUS_LIST = '"Not started,In progress,Done"'
CONFIDENCE_LIST = '"1 Shaky,2 Patchy,3 OK,4 Solid,5 Exam ready"'

DATE_FMT = "dd mmm yyyy"

PLAN_FIRST_ROW = 5
PLAN_HEADER_ROW = 4

# --------------------------------------------------------------------------
# Style helpers.
# --------------------------------------------------------------------------
thin_grid = Side(style="thin", color=GRID)
rule_line = Side(style="thin", color=RULE)
amber_side = Side(style="medium", color=AMBER)

row_border = Border(bottom=thin_grid)
input_border = Border(left=amber_side, right=amber_side, top=amber_side, bottom=amber_side)


def banner(ws, row, last_col, title, subtitle=None):
    """Navy title bar with white text, and an optional grey subtitle beneath."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, last_col + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[row].height = 34
    if subtitle is not None:
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=last_col)
        sub = ws.cell(row=row + 1, column=1, value=subtitle)
        sub.font = Font(name=FONT_NAME, size=10, color=SUBTITLE)
        sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row + 1].height = 20


def section_head(ws, row, last_col, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
    for col in range(1, last_col + 1):
        ws.cell(row=row, column=col).border = Border(bottom=rule_line)
    ws.row_dimensions[row].height = 22


def table_header(ws, row, headers):
    for idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=text)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def print_setup(ws, landscape, title_rows=None):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    if title_rows:
        ws.print_title_rows = title_rows


# --------------------------------------------------------------------------
# Plan rows.
# --------------------------------------------------------------------------
def build_plan_rows():
    """One row per chapter, in course then chapter order."""
    week_of = {}
    for week, (course, chapters, _focus) in WEEKS.items():
        for chapter in chapters:
            week_of[(course, chapter)] = week
    week_of.update(UNSCHEDULED)

    focus_of = {week: focus for week, (_c, _ch, focus) in WEEKS.items()}
    focus_of.update(EXTRA_WEEKS)

    rows = []
    for course in ("RH124", "RH134"):
        for chapter in sorted(CHAPTERS[course]):
            week = week_of[(course, chapter)]
            rows.append(
                {
                    "course": course,
                    "chapter": chapter,
                    "title": CHAPTERS[course][chapter],
                    "week": week,
                    "focus": focus_of[week],
                    "note": NOTES.get((course, chapter), ""),
                }
            )
    return rows


# --------------------------------------------------------------------------
# Sheets.
# --------------------------------------------------------------------------
def build_start_here(wb):
    ws = wb.create_sheet("Start here")
    ws.sheet_properties.tabColor = NAVY
    ws.sheet_view.showGridLines = False
    set_widths(ws, [38, 16, 2, 15, 14, 62])

    banner(
        ws,
        1,
        6,
        "ICT257 study planner",
        "Red Hat System Administration on RHEL 10. Fifteen weeks, 41 chapters, one RHCSA (EX200) exam at the end.",
    )

    body = Font(name=FONT_NAME, size=10, color=NAVY)
    note = Font(name=FONT_NAME, size=10, color=SUBTITLE)
    value = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)

    section_head(ws, 4, 6, "1. Set one date")

    labels = [
        (5, "Date of your week 1 class", None,
         "This is the only cell you have to fill in. Every date in this workbook is worked out from it."),
        (6, "Exam date, week 15", "=SemesterStart+7*14",
         "The same weekday as your class, fourteen weeks on. Type your booked date over it if you have one."),
        (7, "Last day to revise", "=ExamDate-2",
         "Nothing is scheduled after this day. The two days before the exam are yours."),
        (8, "Days until the exam", "=ExamDate-TODAY()",
         "Counts down on its own every time you open the file."),
    ]
    for row, label, formula, hint in labels:
        ws.cell(row=row, column=1, value=label).font = body
        cell = ws.cell(row=row, column=2)
        if formula is None:
            cell.value = DEFAULT_START
        else:
            cell.value = formula
        cell.font = value
        cell.number_format = DATE_FMT if row != 8 else "0"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        hint_cell = ws.cell(row=row, column=6, value=hint)
        hint_cell.font = note
        hint_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22

    ws["B5"].border = input_border
    ws["B5"].fill = PatternFill("solid", fgColor=WHITE)
    ws["D5"] = "Change me"
    ws["D5"].font = Font(name=FONT_NAME, size=9, bold=True, color=AMBER)
    ws["D5"].alignment = Alignment(horizontal="left", vertical="center")

    section_head(ws, 10, 6, "2. Where you are")

    metrics = [
        (11, "Chapters done, out of 41", '=COUNTIF(PlanStatus,"Done")', "=B11/41"),
        (12, "RH124 chapters done, out of 21", '=COUNTIFS(PlanCourse,"RH124",PlanStatus,"Done")', "=B12/21"),
        (13, "RH134 chapters done, out of 20", '=COUNTIFS(PlanCourse,"RH134",PlanStatus,"Done")', "=B13/20"),
        (14, "Practice challenges done, out of 22", '=COUNTIF(PracticeStatus,"Done")', "=B14/22"),
    ]
    for row, label, count_formula, pct_formula in metrics:
        ws.cell(row=row, column=1, value=label).font = body
        count = ws.cell(row=row, column=2, value=count_formula)
        count.font = value
        count.alignment = Alignment(horizontal="center", vertical="center")
        bar = ws.cell(row=row, column=5, value=pct_formula)
        bar.font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
        bar.number_format = "0%"
        bar.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
    ws.cell(row=10, column=5, value="Progress").font = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)

    ws.conditional_formatting.add(
        "E11:E14",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=BLUE_3, showValue=True),
    )

    overdue = (
        '=SUMPRODUCT((PlanFirstPass<>"")*(PlanFirstPass<TODAY())*(PlanStatus<>"Done"))'
        '+SUMPRODUCT((PlanRev1<>"")*(PlanRev1<TODAY())*(PlanStatus<>"Done"))'
        '+SUMPRODUCT((PlanRev2<>"")*(PlanRev2<TODAY())*(PlanStatus<>"Done"))'
    )
    soon = (
        '=SUMPRODUCT((PlanFirstPass<>"")*(PlanFirstPass>=TODAY())*(PlanFirstPass<=TODAY()+7))'
        '+SUMPRODUCT((PlanRev1<>"")*(PlanRev1>=TODAY())*(PlanRev1<=TODAY()+7))'
        '+SUMPRODUCT((PlanRev2<>"")*(PlanRev2>=TODAY())*(PlanRev2<=TODAY()+7))'
    )
    shaky = '=COUNTIF(PlanConfidence,"1*")+COUNTIF(PlanConfidence,"2*")'

    attention = [
        (16, "Dates gone by with the work not done", overdue,
         "Anything above zero is behind. Clear the oldest first."),
        (17, "Dates falling in the next seven days", soon,
         "This is your week. Sort the Study plan by any date column to see them."),
        (18, "Chapters you rated shaky or patchy", shaky,
         "Revise these before anything you already rated solid."),
    ]
    for row, label, formula, hint in attention:
        ws.cell(row=row, column=1, value=label).font = body
        cell = ws.cell(row=row, column=2, value=formula)
        cell.font = value
        cell.number_format = "0"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        hint_cell = ws.cell(row=row, column=6, value=hint)
        hint_cell.font = note
        hint_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 20

    ws.conditional_formatting.add(
        "B16:B18",
        FormulaRule(
            formula=["B16>0"],
            fill=PatternFill("solid", bgColor=AMBER),
            font=Font(name=FONT_NAME, size=11, bold=True, color=NAVY),
        ),
    )

    section_head(ws, 20, 6, "3. How to use it")

    how_to = [
        "Set the date of your week 1 class above. Every other date follows from it.",
        "Go to the Study plan sheet. Each chapter carries three dates. The class date is the week it is taught. Revision 1 is a week later. Revision 2 is three weeks after that.",
        "Three short passes beat one long one. The gap is the point. You have to work to recall the material, and the effort is what fixes it.",
        "Mark Status from the dropdown as you go. Rate your confidence from the dropdown after each revision.",
        "Then revise by weakness, not by order. Filter Confidence to 1 Shaky and 2 Patchy and start there.",
        "Amber means a date has gone by and the work is not done. Light blue means a date falls in the next seven days.",
        "Use the arrows in the header row to filter and sort. Sort by Revision 1 to see what is next.",
        "The Practice sheet holds 22 optional challenges. Each one shows the week it becomes possible. Do the chapter labs first.",
        "Print in landscape. The header row repeats on every page.",
    ]
    row = 21
    for line in how_to:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = Font(name=FONT_NAME, size=10, color=NAVY)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 15 if len(line) < 95 else 28
        row += 1

    row += 1
    section_head(ws, row, 6, "4. How the dates are worked out, and where they had to bend")
    row += 1
    reasoning = [
        "Weeks 1 to 6 are RH124. Weeks 7 to 12 are RH134. Week 13 is catch-up, week 14 is revision, and you sit the exam in week 15.",
        "The two comprehensive reviews sit in weeks 13 and 14, where lessons.md puts them. RH124 20 is open to you from week 6 and RH134 19.02 to 19.04 from week 11, so start either one early if you can.",
        "A comprehensive review gets one revision, not two. There is no room for a second before the exam, and the exam itself is the second pass.",
        "The two preface chapters get no revision dates. They orient you to the courseware and teach no skill the exam asks for. RH134 00 is not scheduled at all.",
        "Three weeks after the first revision falls past the exam for the week 11 and week 12 chapters. Those second revisions are pulled back into revision week instead, week 11 material first and week 12 material a day later.",
        "That squeeze lands on SELinux, storage, booting, troubleshooting, firewalls and network file systems, which is the hardest material in the module. Give revision week to those and start them before the sheet tells you to.",
    ]
    for line in reasoning:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = Font(name=FONT_NAME, size=10, color=NAVY)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 15 if len(line) < 95 else 28
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    credit = ws.cell(
        row=row,
        column=1,
        value="ICT257 Red Hat System Administration, Singapore University of Social Sciences. Taught by Eugene Teo. "
              "Chapters and weeks follow lessons.md. Rebuild this file with python3 planner/build.py.",
    )
    credit.font = Font(name=FONT_NAME, size=9, italic=True, color=SUBTITLE)
    credit.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 26

    print_setup(ws, landscape=True)
    return ws


def build_plan(wb, rows):
    ws = wb.create_sheet("Study plan")
    ws.sheet_properties.tabColor = BLUE_1
    ws.sheet_view.showGridLines = False
    set_widths(ws, [9, 5, 46, 6, 38, 13, 13, 13, 14, 14, 34])

    banner(
        ws,
        1,
        11,
        "Study plan",
        None,
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    sub = ws.cell(
        row=2,
        column=1,
        value='="Every chapter of both courses, in the week it is taught. Your exam is on "'
              '&TEXT(ExamDate,"dd mmm yyyy")&", which is "&ExamDate-TODAY()&" days away."',
    )
    sub.font = Font(name=FONT_NAME, size=10, color=SUBTITLE)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8

    headers = [
        "Course", "Ch", "Chapter", "Week", "What the week covers",
        "Class date", "Revision 1", "Revision 2", "Status", "Confidence", "Notes",
    ]
    table_header(ws, PLAN_HEADER_ROW, headers)

    text_font = Font(name=FONT_NAME, size=10, color=NAVY)
    quiet_font = Font(name=FONT_NAME, size=9, color=SUBTITLE)
    course_font = Font(name=FONT_NAME, size=10, bold=True, color=BLUE_1)

    for offset, item in enumerate(rows):
        r = PLAN_FIRST_ROW + offset

        ws.cell(row=r, column=1, value=item["course"]).font = course_font
        ws.cell(row=r, column=2, value=item["chapter"]).font = text_font
        ws.cell(row=r, column=3, value=item["title"]).font = text_font
        ws.cell(row=r, column=4, value=item["week"]).font = text_font
        ws.cell(row=r, column=5, value=item["focus"]).font = quiet_font

        ws.cell(row=r, column=6, value=f"=SemesterStart+7*($D{r}-1)")
        ws.cell(row=r, column=7, value=f'=IF($B{r}=0,"",MIN($F{r}+7,RevisionCutoff))')
        ws.cell(row=r, column=8, value=f'=IF(OR($B{r}=0,$D{r}>=13),"",MIN($G{r}+21,RevisionCutoff-12+$D{r}))')

        ws.cell(row=r, column=9, value="Not started").font = text_font
        ws.cell(row=r, column=10).font = text_font
        ws.cell(row=r, column=11, value=item["note"]).font = quiet_font

        for col in range(6, 9):
            cell = ws.cell(row=r, column=col)
            cell.font = text_font
            cell.number_format = DATE_FMT

        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.border = row_border
            if col in (2, 4, 6, 7, 8):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col in (3, 5, 11):
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top")

        if offset > 0 and rows[offset - 1]["course"] != item["course"]:
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = Border(top=Side(style="thin", color=RULE), bottom=thin_grid)

    last_row = PLAN_FIRST_ROW + len(rows) - 1

    ws.freeze_panes = ws.cell(row=PLAN_FIRST_ROW, column=3)
    ws.auto_filter.ref = f"A{PLAN_HEADER_ROW}:K{last_row}"

    status_dv = DataValidation(type="list", formula1=STATUS_LIST, allow_blank=True, showErrorMessage=True)
    status_dv.error = "Pick one of the three values from the list."
    status_dv.errorTitle = "Use the dropdown"
    ws.add_data_validation(status_dv)
    status_dv.add(f"I{PLAN_FIRST_ROW}:I{last_row}")

    conf_dv = DataValidation(type="list", formula1=CONFIDENCE_LIST, allow_blank=True, showErrorMessage=True)
    conf_dv.error = "Pick one of the five values from the list."
    conf_dv.errorTitle = "Use the dropdown"
    ws.add_data_validation(conf_dv)
    conf_dv.add(f"J{PLAN_FIRST_ROW}:J{last_row}")

    date_range = f"F{PLAN_FIRST_ROW}:H{last_row}"
    ws.conditional_formatting.add(
        date_range,
        FormulaRule(
            formula=[f'AND(ISNUMBER(F{PLAN_FIRST_ROW}),F{PLAN_FIRST_ROW}<TODAY(),$I{PLAN_FIRST_ROW}<>"Done")'],
            fill=PatternFill("solid", bgColor=AMBER),
            font=Font(name=FONT_NAME, size=10, bold=True, color=NAVY),
        ),
    )
    ws.conditional_formatting.add(
        date_range,
        FormulaRule(
            formula=[
                f'AND(ISNUMBER(F{PLAN_FIRST_ROW}),F{PLAN_FIRST_ROW}>=TODAY(),'
                f'F{PLAN_FIRST_ROW}<=TODAY()+7,$I{PLAN_FIRST_ROW}<>"Done")'
            ],
            fill=PatternFill("solid", bgColor=BLUE_4),
        ),
    )
    ws.conditional_formatting.add(
        f"A{PLAN_FIRST_ROW}:K{last_row}",
        FormulaRule(
            formula=[f'$I{PLAN_FIRST_ROW}="Done"'],
            font=Font(name=FONT_NAME, size=10, color=SUBTITLE, strike=True),
        ),
    )
    ws.conditional_formatting.add(
        f"J{PLAN_FIRST_ROW}:J{last_row}",
        FormulaRule(
            formula=[f'OR($J{PLAN_FIRST_ROW}="1 Shaky",$J{PLAN_FIRST_ROW}="2 Patchy")'],
            font=Font(name=FONT_NAME, size=10, bold=True, color=AMBER),
        ),
    )

    print_setup(ws, landscape=True, title_rows=f"{PLAN_HEADER_ROW}:{PLAN_HEADER_ROW}")
    return ws, last_row


def build_practice(wb):
    ws = wb.create_sheet("Practice")
    ws.sheet_properties.tabColor = BLUE_2
    ws.sheet_view.showGridLines = False
    set_widths(ws, [5, 50, 8, 13, 11, 34, 28, 14, 30])

    banner(ws, 1, 9, "Practice challenges", None)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    sub = ws.cell(
        row=2,
        column=1,
        value="Optional. None of it is required for ICT257. Do the chapter labs and the guided exercises first, "
              "then come here to find out whether you can reach a result with nothing open in front of you.",
    )
    sub.font = Font(name=FONT_NAME, size=10, color=SUBTITLE)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8

    headers = ["#", "Challenge", "Ready in", "Ready date", "Difficulty", "Objectives", "Prepare with", "Status", "Notes"]
    table_header(ws, 4, headers)

    text_font = Font(name=FONT_NAME, size=10, color=NAVY)
    quiet_font = Font(name=FONT_NAME, size=9, color=SUBTITLE)
    mono_font = Font(name="Consolas", size=9, color=BLUE_1)

    first = 5
    for offset, (num, title, week, difficulty, objectives, prepare) in enumerate(PRACTICE):
        r = first + offset
        ws.cell(row=r, column=1, value=num).font = text_font
        ws.cell(row=r, column=2, value=title).font = text_font
        ws.cell(row=r, column=3, value=f"week {week}").font = quiet_font
        ws.cell(row=r, column=4, value=f"=SemesterStart+7*({week}-1)")
        ws.cell(row=r, column=4).number_format = DATE_FMT
        ws.cell(row=r, column=4).font = text_font
        ws.cell(row=r, column=5, value=difficulty).font = text_font
        ws.cell(row=r, column=6, value=objectives).font = quiet_font
        ws.cell(row=r, column=7, value=prepare).font = mono_font
        ws.cell(row=r, column=8, value="Not started").font = text_font
        ws.cell(row=r, column=9).font = quiet_font

        for col in range(1, 10):
            cell = ws.cell(row=r, column=col)
            cell.border = row_border
            if col in (1, 3, 4, 5):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    last_row = first + len(PRACTICE) - 1

    ws.freeze_panes = ws.cell(row=first, column=3)
    ws.auto_filter.ref = f"A4:I{last_row}"

    status_dv = DataValidation(type="list", formula1=STATUS_LIST, allow_blank=True, showErrorMessage=True)
    status_dv.error = "Pick one of the three values from the list."
    status_dv.errorTitle = "Use the dropdown"
    ws.add_data_validation(status_dv)
    status_dv.add(f"H{first}:H{last_row}")

    ws.conditional_formatting.add(
        f"D{first}:D{last_row}",
        FormulaRule(
            formula=[f'AND(ISNUMBER($D{first}),$D{first}<=TODAY(),$H{first}<>"Done")'],
            fill=PatternFill("solid", bgColor=BLUE_4),
        ),
    )
    ws.conditional_formatting.add(
        f"A{first}:I{last_row}",
        FormulaRule(
            formula=[f'$H{first}="Done"'],
            font=Font(name=FONT_NAME, size=10, color=SUBTITLE, strike=True),
        ),
    )

    print_setup(ws, landscape=True, title_rows="4:4")
    return ws, last_row


def add_names(wb, plan_last_row, practice_last_row):
    names = {
        "SemesterStart": "'Start here'!$B$5",
        "ExamDate": "'Start here'!$B$6",
        "RevisionCutoff": "'Start here'!$B$7",
        "PlanCourse": f"'Study plan'!$A${PLAN_FIRST_ROW}:$A${plan_last_row}",
        "PlanFirstPass": f"'Study plan'!$F${PLAN_FIRST_ROW}:$F${plan_last_row}",
        "PlanRev1": f"'Study plan'!$G${PLAN_FIRST_ROW}:$G${plan_last_row}",
        "PlanRev2": f"'Study plan'!$H${PLAN_FIRST_ROW}:$H${plan_last_row}",
        "PlanStatus": f"'Study plan'!$I${PLAN_FIRST_ROW}:$I${plan_last_row}",
        "PlanConfidence": f"'Study plan'!$J${PLAN_FIRST_ROW}:$J${plan_last_row}",
        "PracticeStatus": f"'Practice'!$H$5:$H${practice_last_row}",
        "PracticeReady": f"'Practice'!$D$5:$D${practice_last_row}",
    }
    for name, ref in names.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))


def main():
    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "ict257-study-planner.xlsx")

    rows = build_plan_rows()
    assert len(rows) == 41, f"expected 41 chapters, got {len(rows)}"

    wb = Workbook()
    wb.remove(wb.active)

    build_start_here(wb)
    _plan_ws, plan_last_row = build_plan(wb, rows)
    _practice_ws, practice_last_row = build_practice(wb)
    add_names(wb, plan_last_row, practice_last_row)

    wb.properties.title = "ICT257 study planner"
    wb.properties.subject = "Red Hat System Administration, RHCSA (EX200) preparation"
    wb.properties.creator = "ICT257, Singapore University of Social Sciences"

    wb.active = 0
    wb.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
